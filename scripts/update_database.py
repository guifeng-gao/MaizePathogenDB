#!/usr/bin/env python3
"""
MaizePathogenDB Master Update Pipeline.

Run this after updating 玉米细菌性病原菌数据库.xlsx, 玉米病毒性病原菌数据库.xlsx,
or 玉米真菌性病原菌数据库.xlsx (adding new species).

Usage:
    python3 update_database.py

Workflow:
  1. Parse all 3 Excel files → extract species list
  2. Build/update all_taxids.json
  3. Download marker gene sequences from NCBI
  4. Write FASTA files (all + per-category)
  5. Build BLAST databases
  6. Generate web platform data (taxonomy.json, seq_meta.json, fingerprints.json)
  7. Generate all figures
  8. Update handoff.md stats
  9. Re-run NCBI-nt comparison (if new species added)
"""

import os, json, sys, subprocess, time, requests, re
from collections import defaultdict
import openpyxl

# ── Config ──────────────────────────────────────────────────────────
BASE = "/Users/gfgao/Desktop/blacksoil_metaG"
DB_DIR = os.path.join(BASE, "maize_pathogen_db")
SEQ_DIR = os.path.join(DB_DIR, "sequences")
TAX_DIR = os.path.join(DB_DIR, "taxonomy")
BLAST_DIR = os.path.join(DB_DIR, "blast_db")
DOCS_DIR = os.path.join(DB_DIR, "docs")
VAL_DIR = os.path.join(DOCS_DIR, "validation")
WEB_DIR = os.path.join(BASE, "maize_pathogen_web")
WEB_DATA_DIR = os.path.join(WEB_DIR, "data")
SCRIPT_DIR = os.path.join(DB_DIR, "scripts")
HANDOFF = os.path.join(BASE, "handoff.md")

EXCEL_FILES = {
    "bacteria": os.path.join(BASE, "玉米细菌性病原菌数据库.xlsx"),
    "viruses": os.path.join(BASE, "玉米病毒性病原菌数据库.xlsx"),
    "fungi": os.path.join(BASE, "玉米真菌性病原菌数据库.xlsx"),
}

ENTREZ_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
EMAIL = "maize_pathogen_db@example.com"

def asciify(s):
    """Remove non-ASCII characters from a string (for BLAST compatibility)."""
    return re.sub(r'[^\x20-\x7E]', '_', s)

# ── Excel Column Maps ──────────────────────────────────────────────
# (cat) -> (name_col, disease_cn_col, taxid_col, taxonomy_cols)
# Columns are 1-indexed; taxonomy_cols = (kingdom, phylum, class, order, family, genus, species)
EXCEL_SCHEMA = {
    "bacteria": {
        "name_col": 2,          # 病原菌学名
        "disease_cn_col": 4,    # 病害中文名
        "taxid_col": 12,        # NCBI TaxID
        "taxonomy_cols": (5, 6, 7, 8, 9, 10, 11),  # 界门纲目科属种
        "data_start": 4,        # data starts at row 4
    },
    "viruses": {
        "name_col": 2,          # 病毒名称
        "disease_cn_col": 4,    # 病害中文名
        "taxid_col": 12,        # NCBI TaxID
        "taxonomy_cols": (5, 6, 7, 8, 9, 10, 11),  # 域门纲目科属种
        "data_start": 4,
    },
    "fungi": {
        "name_col": 3,          # 病原菌学名 (scientific name)
        "disease_cn_col": 5,    # 病害中文名
        "taxid_col": 13,        # NCBI TaxID
        "taxonomy_cols": (6, 7, 8, 9, 10, 11, 12),  # 界门纲目科属种
        "data_start": 4,
    },
}

os.makedirs(SEQ_DIR, exist_ok=True)
os.makedirs(os.path.join(SEQ_DIR, "bacteria"), exist_ok=True)
os.makedirs(os.path.join(SEQ_DIR, "viruses"), exist_ok=True)
os.makedirs(os.path.join(SEQ_DIR, "fungi"), exist_ok=True)
os.makedirs(BLAST_DIR, exist_ok=True)
os.makedirs(TAX_DIR, exist_ok=True)
os.makedirs(WEB_DATA_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════
def log(msg):
    print(f"[update] {msg}", flush=True)

# ═══════════════════════════════════════════════════════════════════
# STEP 1: Parse Excel Files
# ═══════════════════════════════════════════════════════════════════
log("=" * 60)
log("STEP 1/8: Parsing Excel species files")
log("=" * 60)

def extract_taxid(raw):
    """Extract TaxID from various formats: 'TaxID=XXX', 'XXX', 'ssp TaxID=XXX'"""
    if raw is None:
        return None
    s = str(raw).strip()
    if "TaxID=" in s:
        s = s.split("TaxID=")[1]
    # Take first part (might be "Fusarium verticillioides" as taxid in fungi)
    s = s.split()[0].strip()
    if s.isdigit():
        return s
    return None

species_records = []
excel_stats = {"bacteria": 0, "viruses": 0, "fungi": 0}

for cat, fpath in EXCEL_FILES.items():
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    schema = EXCEL_SCHEMA[cat]
    
    n_species = 0
    for r in range(schema["data_start"], ws.max_row + 1):
        name = ws.cell(r, schema["name_col"]).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name.startswith("TaxID="):
            continue
        
        disease_cn = str(ws.cell(r, schema["disease_cn_col"]).value or "").strip()
        raw_taxid = ws.cell(r, schema["taxid_col"]).value
        taxid = extract_taxid(raw_taxid)
        if not taxid:
            log(f"  ⚠ No TaxID for row {r}: {name}")
            continue
        
        # Get disease_en from different column depending on category
        disease_en = str(ws.cell(r, 3).value or "").strip()
        
        # Parse taxonomy lineage
        kc, pc, cc, oc, fc, gc, sc = schema["taxonomy_cols"]
        tax = {
            "kingdom": str(ws.cell(r, kc).value or "").strip(),
            "phylum": str(ws.cell(r, pc).value or "").strip(),
            "class": str(ws.cell(r, cc).value or "").strip(),
            "order": str(ws.cell(r, oc).value or "").strip(),
            "family": str(ws.cell(r, fc).value or "").strip(),
            "genus": str(ws.cell(r, gc).value or "").strip(),
            "species": str(ws.cell(r, sc).value or "").strip(),
        }
        
        species_records.append({
            "tax_id": taxid,
            "species": name.replace('\n', ' ').replace('\r', ' '),
            "category": cat,
            "disease_cn": disease_cn,
            "disease_en": disease_en,
            "taxonomy": tax,
        })
        n_species += 1
    
    excel_stats[cat] = n_species
    log(f"  {cat}: {n_species} species parsed from Excel")

total_species = sum(excel_stats.values())
log(f"  Total: {total_species} species")

# Write all_taxids.json
taxids_path = os.path.join(TAX_DIR, "all_taxids.json")
simple_records = [{
    "tax_id": r["tax_id"], "species": r["species"],
    "category": r["category"], "disease_cn": r["disease_cn"]
} for r in species_records]
with open(taxids_path, "w") as f:
    json.dump(simple_records, f, indent=2, ensure_ascii=False)
log(f"  ✓ Saved: {taxids_path}")

# Write per-category taxid files
for cat in ["bacteria", "viruses", "fungi"]:
    taxids = [r["tax_id"] for r in species_records if r["category"] == cat]
    taxid_path = os.path.join(TAX_DIR, f"taxids_{cat}.txt")
    with open(taxid_path, "w") as f:
        f.write("\n".join(taxids))
    log(f"  ✓ Saved: {taxid_path}")

# ═══════════════════════════════════════════════════════════════════
# STEP 2: Download Sequences
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 2/8: Downloading marker gene sequences from NCBI")
log("=" * 60)

# Download sequences for each species
def esearch(db, term, retmax=5):
    params = {"db": db, "term": term, "retmax": retmax, "retmode": "json",
              "email": EMAIL, "tool": "maize_pathogen_db", "sort": "sequence_length"}
    r = requests.get(f"{ENTREZ_BASE}/esearch.fcgi", params=params, timeout=30)
    if r.status_code == 429:
        time.sleep(5)
        r = requests.get(f"{ENTREZ_BASE}/esearch.fcgi", params=params, timeout=30)
    r.raise_for_status()
    return r.json().get("esearchresult", {}).get("idlist", []), \
           int(r.json().get("esearchresult", {}).get("count", 0))

def efetch(db, ids):
    params = {"db": db, "id": ",".join(ids), "rettype": "fasta", "retmode": "text",
              "email": EMAIL, "tool": "maize_pathogen_db"}
    r = requests.get(f"{ENTREZ_BASE}/efetch.fcgi", params=params, timeout=60)
    if r.status_code == 429:
        time.sleep(5)
        r = requests.get(f"{ENTREZ_BASE}/efetch.fcgi", params=params, timeout=60)
    r.raise_for_status()
    return r.text

all_seqs = {}
download_stats = {"bacteria": 0, "viruses": 0, "fungi": 0}
downloaded_taxids = set()

for i, rec in enumerate(species_records):
    taxid = rec["tax_id"]
    sp = rec["species"]
    cat = rec["category"]
    disease = rec["disease_cn"]
    
    log(f"  [{i+1}/{total_species}] {cat}: {sp[:50]} (txid{taxid})")
    
    if cat == "bacteria":
        query = (f'txid{taxid}[Organism] AND ("16S ribosomal RNA"[Gene] OR "16S rRNA"[Title]) AND 1000:2000[SLEN]')
    elif cat == "fungi":
        query = (f'txid{taxid}[Organism] AND ("internal transcribed spacer"[Title] OR "ITS1"[Title] '
                 f'OR "ITS2"[Title] OR "5.8S"[Title] OR "18S ribosomal RNA"[Title]) AND 400:2000[SLEN]')
    elif cat == "viruses":
        query = (f'txid{taxid}[Organism] AND ("complete genome"[Title] OR "segment"[Title] '
                 f'OR "polyprotein"[Gene] OR "coat protein"[Gene]) '
                 f'NOT ("clone"[Title] OR "vector"[Title] OR "synthetic"[Title])')
    else:
        continue
    
    try:
        ids, count = esearch("nucleotide", query, retmax=3)
        if ids:
            fasta_text = efetch("nucleotide", ids)
            n_seqs = 0
            for block in fasta_text.strip().split("\n\n"):
                lines = block.strip().split("\n")
                if len(lines) < 2:
                    continue
                hdr = lines[0].strip()
                seq = "".join(l.strip() for l in lines[1:])
                if len(seq) >= 100:
                    clean_hdr = f">{taxid}|{sp[:80]}|{cat}|{disease[:80]}"
                    full_hdr = f"{clean_hdr} {hdr[1:]}"
                    all_seqs[full_hdr] = seq
                    n_seqs += 1
                    downloaded_taxids.add(taxid)
            
            if n_seqs > 0:
                download_stats[cat] += 1
    except Exception as e:
        log(f"    ✗ ERROR: {str(e)[:80]}")
    
    time.sleep(0.5)

log(f"\n  Download summary:")
log(f"  Total sequences: {len(all_seqs)}")
log(f"  Species covered: {len(downloaded_taxids)}/{total_species}")
for cat in ["bacteria", "viruses", "fungi"]:
    n_tax = len([r for r in species_records if r["category"] == cat])
    log(f"    {cat}: {download_stats[cat]}/{n_tax}")

# ═══════════════════════════════════════════════════════════════════
# STEP 3: Write FASTA Files
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 3/8: Writing FASTA files")
log("=" * 60)

def write_fasta(fpath, seqs_dict):
    os.makedirs(os.path.dirname(fpath), exist_ok=True)
    with open(fpath, "w") as f:
        for hdr, seq in seqs_dict.items():
            f.write(hdr + "\n")
            for j in range(0, len(seq), 80):
                f.write(seq[j:j+80] + "\n")

# All sequences
all_fasta = os.path.join(SEQ_DIR, "maize_pathogens_all.fasta")
write_fasta(all_fasta, all_seqs)
log(f"  ✓ {all_fasta} ({len(all_seqs)} seqs)")

# Per-category
for cat in ["bacteria", "viruses", "fungi"]:
    cat_seqs = {h: s for h, s in all_seqs.items() if f"|{cat}|" in h}
    cat_path = os.path.join(SEQ_DIR, cat, f"maize_pathogens_{cat}.fasta")
    write_fasta(cat_path, cat_seqs)
    log(f"  ✓ {cat_path} ({len(cat_seqs)} seqs)")

# ═══════════════════════════════════════════════════════════════════
# STEP 4: Build BLAST Databases
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 4/8: Building BLAST databases")
log("=" * 60)

# Find makeblastdb
BLAST_BIN = "/tmp/ncbi-blast-2.17.0+/bin/makeblastdb"
if not os.path.exists(BLAST_BIN):
    for p in [os.path.expanduser("~/anaconda3/bin/makeblastdb"),
              os.path.expanduser("~/miniconda3/bin/makeblastdb"),
              "/usr/local/bin/makeblastdb", "/opt/homebrew/bin/makeblastdb"]:
        if os.path.exists(p):
            BLAST_BIN = p
            break

blast_databases = [
    ("maize_pathogens_all", os.path.join(SEQ_DIR, "maize_pathogens_all.fasta")),
    ("maize_pathogens_bacteria", os.path.join(SEQ_DIR, "bacteria", "maize_pathogens_bacteria.fasta")),
    ("maize_pathogens_viruses", os.path.join(SEQ_DIR, "viruses", "maize_pathogens_viruses.fasta")),
    ("maize_pathogens_fungi", os.path.join(SEQ_DIR, "fungi", "maize_pathogens_fungi.fasta")),
]

for db_name, fasta_path in blast_databases:
    if not os.path.exists(fasta_path):
        log(f"  ⚠ Skipping {db_name}: {fasta_path} not found")
        continue
    db_out = os.path.join(BLAST_DIR, db_name)
    cmd = [BLAST_BIN, "-in", fasta_path, "-dbtype", "nucl", "-out", db_out,
           "-title", f"Maize Pathogen Database - {db_name}"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode == 0:
        log(f"  ✓ {db_name} -> {db_out}")
    else:
        log(f"  ✗ {db_name}: {result.stderr[:100]}")

# ═══════════════════════════════════════════════════════════════════
# STEP 5: Generate Web Data
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 5/8: Generating web platform data")
log("=" * 60)

# 5a. taxonomy.json — species catalog for Species Search
taxonomy_list = []
cat_map = {"bacteria": "Bacteria", "viruses": "Virus", "fungi": "Fungi"}
for rec in species_records:
    tax = rec["taxonomy"]
    disease_en = rec.get("disease_en", "")
    # Properly separate fungi vs oomycetes
    phylum = tax.get("phylum", "")
    from_cat = rec["category"]
    if from_cat == "fungi" and ("Oomycota" in phylum or "卵菌" in phylum):
        display_cat = "Oomycetes"
    else:
        display_cat = {"bacteria": "Bacteria", "viruses": "Virus", "fungi": "Fungi"}.get(from_cat, from_cat)
    
    keywords = f"{rec['species'].lower()} " + " ".join(
        v.lower() for v in tax.values() if v
    ) + f" {disease_en.lower()} {rec['disease_cn']}"
    taxonomy_list.append({
        "species": rec["species"],
        "disease_en": disease_en,
        "disease_cn": rec["disease_cn"],
        "kingdom": tax["kingdom"],
        "phylum": tax["phylum"],
        "class": tax["class"],
        "order": tax["order"],
        "family": tax["family"],
        "genus": tax["genus"],
        "taxid": rec["tax_id"],
        "category": display_cat,
        "keywords": keywords,
    })

taxonomy_path = os.path.join(WEB_DATA_DIR, "taxonomy.json")
with open(taxonomy_path, "w") as f:
    json.dump(taxonomy_list, f, indent=2, ensure_ascii=False)
log(f"  ✓ taxonomy.json ({len(taxonomy_list)} species)")

# 5b. seq_meta.json — sequence length metadata for Sequence Search
seq_meta = {}
for hdr, seq in all_seqs.items():
    parts = hdr.split("|")
    taxid = parts[0].replace(">", "")
    species = parts[1]
    cat = parts[2]
    seq_meta[taxid] = {"s": species, "c": cat, "l": len(seq)}

seq_meta_path = os.path.join(WEB_DATA_DIR, "seq_meta.json")
with open(seq_meta_path, "w") as f:
    json.dump(seq_meta, f, indent=2, ensure_ascii=False)
log(f"  ✓ seq_meta.json ({len(seq_meta)} entries)")

# 5c. fingerprints.json — 25-mer fingerprints for Sequence Search
def extract_kmers(seq, k=25, n=50):
    """Extract n evenly distributed k-mers from a sequence."""
    if len(seq) < k:
        return [seq]
    step = max(1, (len(seq) - k) // max(n, 1))
    return [seq[i:i+k] for i in range(0, len(seq) - k + 1, step)][:n]

fingerprints = {}
for hdr, seq in all_seqs.items():
    taxid = hdr.split("|")[0].replace(">", "")
    fingerprints[taxid] = extract_kmers(seq, k=25, n=50)

fingerprints_path = os.path.join(WEB_DATA_DIR, "fingerprints.json")
with open(fingerprints_path, "w") as f:
    json.dump(fingerprints, f, indent=2, ensure_ascii=False)
log(f"  ✓ fingerprints.json ({len(fingerprints)} species fingerprinted)")

# ═══════════════════════════════════════════════════════════════════
# STEP 6: Generate Figures
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 6/8: Generating figures")
log("=" * 60)

fig_script = os.path.join(SCRIPT_DIR, "generate_figures.py")
if os.path.exists(fig_script):
    result = subprocess.run([sys.executable, fig_script], capture_output=True, text=True, timeout=300)
    if result.returncode == 0:
        log(f"  ✓ Figures generated")
        for line in result.stdout.split("\n"):
            if "✓" in line or "ALL" in line:
                log(f"    {line.strip()}")
    else:
        log(f"  ✗ Figure generation failed: {result.stderr[:200]}")

# ═══════════════════════════════════════════════════════════════════
# STEP 7: Re-run NCBI-nt Comparison
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 7/8: Re-running NCBI-nt comparison (if new species added)")
log("=" * 60)

ncbi_script = os.path.join(SCRIPT_DIR, "ncbi_nt_comparison_v2.py")
fig_script2 = os.path.join(SCRIPT_DIR, "generate_ncbi_figure.py")
if os.path.exists(ncbi_script):
    # Clean old comparison cache
    old_json = os.path.join(VAL_DIR, "ncbi_nt_comparison_v2.json")
    if os.path.exists(old_json):
        os.remove(old_json)
    result = subprocess.run([sys.executable, "-u", ncbi_script],
                          capture_output=True, text=True, timeout=3600)
    if result.returncode == 0:
        log(f"  ✓ NCBI-nt comparison completed")
        for line in result.stdout.split("\n"):
            if "Overall" in line or "MaizePathogenDB" in line or "DONE" in line:
                log(f"    {line.strip()}")
        # Generate comparison figure
        if os.path.exists(fig_script2):
            subprocess.run([sys.executable, fig_script2], capture_output=True, text=True, timeout=120)
            log(f"  ✓ NCBI-nt comparison figure generated")
    else:
        log(f"  ✗ NCBI-nt comparison failed: {result.stderr[:200]}")
else:
    log(f"  ⚠ NCBI-nt comparison script not found, skipping")

# ═══════════════════════════════════════════════════════════════════
# STEP 8: Print Update Summary (user updates handoff.md manually)
# ═══════════════════════════════════════════════════════════════════
log(f"\n{'=' * 60}")
log("STEP 8/8: Update Summary")
log("=" * 60)
log(f"\n  Database update complete!")
log(f"  Species parsed: {total_species}")
log(f"  Sequences downloaded: {len(all_seqs)}")
log(f"  Species with sequences: {len(downloaded_taxids)}/{total_species}")
log(f"  Species without sequences: {total_species - len(downloaded_taxids)}")
log(f"\n  Files updated:")
log(f"    {taxids_path}")
log(f"    {all_fasta}")
log(f"    {os.path.join(BLAST_DIR, 'maize_pathogens_all*')}")
log(f"    {taxonomy_path}")
log(f"    {seq_meta_path}")
log(f"    {fingerprints_path}")
log(f"    docs/validation/Fig*.pdf")
log(f"\n  Next steps (manual):")
log(f"    • Update handoff.md stats")
log(f"    • Update web platform index.html data embed")
log(f"    • Commit and push to GitHub")
log(f"\n  Uncovered species (no sequences found):")
for rec in species_records:
    if rec["tax_id"] not in downloaded_taxids:
        log(f"    {rec['tax_id']}: {rec['species'][:50]} ({rec['category']})")
