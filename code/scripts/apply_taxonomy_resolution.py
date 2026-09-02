#!/usr/bin/env python3
"""Apply taxid_resolution.tsv to the canonical Excel and taxonomy JSON files."""

import csv
import json
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESOLUTION = os.path.join(ROOT, "data", "taxid_resolution.tsv")
EXCEL_FILES = [
    os.path.join(ROOT, "Figshare", "Maize Pathogen.xlsx"),
    os.path.join(ROOT, "Maize Pathogen.xlsx"),
]
ALL_TAXIDS = os.path.join(ROOT, "Figshare", "all_taxids.json")
TAXONOMY = os.path.join(ROOT, "Figshare", "taxonomy.json")


def find_excel_row(ws, full_name):
    for row in ws.iter_rows(min_row=2):
        if row[1].value and str(row[1].value).strip() == full_name:
            return row
    return None


def main():
    with open(RESOLUTION, encoding="utf-8") as fh:
        resolution = list(csv.DictReader(fh, delimiter="\t"))

    updates = []
    for item in resolution:
        if item["row_type"] == "duplicate_species_name" and item["status"] == "RESOLVED":
            updates.append((item["original_taxid"], item["original_species"],
                            item["resolved_scientific_name"]))

    wb_paths = []
    for excel_path in EXCEL_FILES:
        if not os.path.exists(excel_path):
            continue
        wb = openpyxl.load_workbook(excel_path)
        changed = 0
        for sheet in wb.worksheets:
            for original_taxid, _, new_species in updates:
                for row in sheet.iter_rows(min_row=2):
                    taxid_cell = row[11]
                    species_cell = row[10]
                    if str(taxid_cell.value).strip() == original_taxid:
                        species_cell.value = new_species
                        changed += 1
        wb.save(excel_path)
        wb_paths.append((excel_path, changed))
        print(f"updated {excel_path}: {changed} species cells")

    all_taxids = json.load(open(ALL_TAXIDS, encoding="utf-8"))
    taxonomy = json.load(open(TAXONOMY, encoding="utf-8"))
    for original_taxid, _, new_species in updates:
        for record in all_taxids:
            if record["tax_id"] == original_taxid:
                record["species"] = new_species
        for record in taxonomy:
            if record.get("taxid") == original_taxid:
                record["species"] = new_species
                record["keywords"] = " ".join([
                    new_species.lower(),
                    (record.get("disease_en") or "").lower(),
                    (record.get("disease_cn") or ""),
                    (record.get("kingdom") or "").lower(),
                    (record.get("phylum") or "").lower(),
                    (record.get("class") or "").lower(),
                    (record.get("order") or "").lower(),
                    (record.get("family") or "").lower(),
                    (record.get("genus") or "").lower(),
                ])

    json.dump(all_taxids, open(ALL_TAXIDS, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(taxonomy, open(TAXONOMY, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"updated {ALL_TAXIDS} and {TAXONOMY}")

    pending = sum(1 for item in resolution
                  if item["row_type"] == "not_verified" and item["status"] == "PENDING")
    print(f"NOT_VERIFIED rows still PENDING: {pending}")


if __name__ == "__main__":
    main()
